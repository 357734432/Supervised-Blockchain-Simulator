from openpyxl import load_workbook

file_path = 'BlockSim-master\(Allverify)1day_1e-06M_0.02K.xlsx'

class readExcel:
    def __init__(self):
        self.wb = load_workbook(file_path)
        self.sheet_names = self.wb.sheetnames

    def read_line(self, sheet, num):
        SHEET = self.wb[self.sheet_names[sheet]]
        name = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'j', 'K']
        max_column = SHEET.max_column
        result = []
        for i in range(1, max_column):
            result.append(SHEET[name[i] + str(num)].value)
        return result

    def get_lines(self, sheet):
        return self.wb[self.sheet_names[sheet]].max_row

